
"""`
   Excelから検索Wordを取得し、それを上から順にGoogleで検索をかけ、
   その一番上に出てきたタイトルと、そのページのtitleタグ中の要素をExcelに書き出し保存する。（スクレイピング）

  採用技術
   言語 Python (HTML)
    技術 スクレイピング クローリング   

  使用上の注意 
   ・プログラムを起動する際にはExcelを開かないようにしてください。
   Excelを開いた状態で起動させると、熱暴走したり、インプットするExcelが壊れてしまうことがあります。

   11/8 検索結果が10件未満の時に次の会社の出力が狂うのを直したい
   11/10検索結果がきれいになった。検証資料を作成し、和司さんに見せる。
      また、赤字になった箇所を数える機能を追加する。
      動画がヒットした時だけ、出力がおかしくなる。（検索ワードそのまま＋「の動画」と出力される）
      見かけが悪いが、反社チェックという意味では問題無いと思われる。
   11/18コードがスパゲッティになってきたのでリファクタリングしたい。
       クエリパラメータを削除するコードを追加したい。
   
   11/24大量のデータを流した結果、何故か途中から最初の会社に戻る現象が生じた。
      原因の究明をするか、或いは少量のデータのみで使用してもらうことにするか、、、。

   12/03どうも90件くらいのところで過重負荷なのか、処理が停止してしまう。

   12/10大体50件刻みでやると効率的である。これ以上だと熱暴走（？）してとまってしまう。

   12/21現時点で認識しているバグを一通り潰せた。
           サイバー攻撃を疑われないようにするためである。

      23/7/25 インプットデータに新しい列が新設される「申請者」「担当者」ので、Excelへの転記の仕方を調節する
   　　　　　また、使用しないコメントアウト行を削除した
   
"""

import requests
from requests.exceptions import ConnectionError, TooManyRedirects, HTTPError
from bs4 import BeautifulSoup
import bs4
import time
import openpyxl
import subprocess
import random
import time
import re

ANTISOCIAL＿FORCES = ' AND (悪徳商法 OR キャッチセールス OR マルチ商法 OR ねずみ講 OR '\
'書類送検 OR 送検 OR 捜査 OR 家宅捜索 OR 捜索 OR 指名手配 OR 容疑者 OR 容疑 OR 被告 OR 釈放 OR '\
'検挙 OR 摘発 OR 犯罪 OR 違反 OR 押収 OR 警察 OR 行政指導 OR 行政処分 OR いたずら電話 OR 違法コピー OR '\
'インサイダー取引 OR 相場操縦 OR 横領 OR 汚職 OR 贈収賄 OR 収賄 OR 贈賄 OR わいろ OR 架空取引 OR '\
'株価操作 OR カラ出張 OR 監禁 OR 機密漏洩 OR 情報漏洩 OR 脅迫 OR 企業脅迫 OR 恐喝 OR 企業犯罪 OR '\
'架空請求 OR 偽造カード OR 偽造キャッシュカード OR スキミング OR ヤミ金融 OR 偽造 OR 通貨偽造 OR '\
'偽造貨幣 OR 偽札 OR 業務上過失致死傷 OR 業務上過失傷害 OR 業務上過失致死 OR 危険運転致死傷 OR '\
'危険運転致死 OR 危険運転致傷 OR 過失運転致死傷 OR 過失運転致死 OR 過失運転致傷 OR 公務執行妨害 OR '\
'コンピューター犯罪 OR サイバーテロ OR ハッカー OR 不正アクセス OR 強盗 OR 銀行強盗 OR 詐欺 OR '\
'オレオレ詐欺 OR 振り込め詐欺 OR フィッシング詐欺 OR 反社 OR 半グレ OR 保険金詐欺 OR 死体遺棄 OR '\
'傷害 OR 人身取引 OR 人身売買 OR ストーカー OR 痴漢 OR 売春 OR 援助交際 OR パパ活 OR 児童買春 OR '\
'児童虐待 OR ＤＶ OR わいせつ OR 児童ポルノ OR 窃盗 OR 空き巣 OR 車上荒らし OR 万引き OR 脱税 OR '\
'申告漏れ OR データ捏造 OR 背任 OR 爆破 OR ひき逃げ OR 放火 OR 密入国 OR 密輸 OR 密漁 OR 誘拐 OR '\
'拉致 OR 罰金 OR 追徴金 OR リンチ OR 暴力団 OR 組長 OR やくざ OR ヤクザ OR 暴力団対策 OR 企業舎弟 OR '\
'フロント企業 OR マフィア OR 構成員 OR 準構成員 OR 組員 OR テロ OR テロリスト OR 仕手筋 OR 麻薬 OR '\
'アヘン OR コカイン OR ヘロイン OR モルヒネ OR 覚せい剤 OR 覚醒剤 OR 大麻 OR えせ OR エセ OR 総会屋 OR '\
'逮捕 OR 告発 OR 起訴 OR 訴訟 OR 殺人 OR 拳銃 OR 銃 OR ピストル OR 発砲 OR 日本刀 OR とばく OR 賭博)'

ANTISOCIAL_FORCES_LIST = ["悪徳商法","キャッチセールス","マルチ商法","ねずみ講"\
    "書類送検","送検","捜査","家宅捜索","捜索","指名手配","容疑者","容疑","被告","釈放"\
    "検挙","摘発","犯罪","違反","押収","警察","行政指導","行政処分","いたずら電話","違法コピー"\
    "インサイダー取引","相場操縦","横領","汚職","贈収賄","収賄","贈賄","わいろ","架空取引"\
    "株価操作","カラ出張","監禁","機密漏洩","情報漏洩","脅迫","企業脅迫","恐喝","企業犯罪"\
    "架空請求","偽造カード","偽造キャッシュカード","スキミング","ヤミ金融","偽造","通貨偽造"\
    "偽造貨幣","偽札","業務上過失致死傷","業務上過失傷害","業務上過失致死","危険運転致死傷"\
    "危険運転致死","危険運転致傷","過失運転致死傷","過失運転致死","過失運転致傷","公務執行妨害"\
    "コンピューター犯罪","サイバーテロ","ハッカー","不正アクセス","強盗","銀行強盗","詐欺"\
    "オレオレ詐欺","振り込め詐欺","フィッシング詐欺","反社","半グレ","保険金詐欺","死体遺棄"\
    "傷害","人身取引","人身売買","ストーカー","痴漢","売春","援助交際","パパ活","児童買春"\
    "児童虐待","ＤＶ","わいせつ","児童ポルノ","窃盗","空き巣","車上荒らし","万引き","脱税"\
    "申告漏れ","データ捏造","背任","爆破","ひき逃げ","放火","密入国","密輸","密漁","誘拐"\
    "拉致","罰金","追徴金","リンチ","暴力団","組長","やくざ","ヤクザ","暴力団対策","企業舎弟"\
    "フロント企業","マフィア","構成員","準構成員","組員","テロ","テロリスト","仕手筋","麻薬"\
    "アヘン","コカイン","ヘロイン","モルヒネ","覚せい剤","覚醒剤","大麻","えせ","エセ","総会屋"\
    "逮捕","告発","起訴","訴訟","殺人","拳銃","銃","ピストル","発砲","日本刀","とばく","賭博"]

#相手サーバに負担をかけないために、タイムスリープを設定（１秒以上取っておくと安全）
#SLEEP_TIME = 1

#2列目以降に検索結果を次々と書き込んでいく
r = 5
# 時間計測開始
time_sta = time.time()
# 上位から何件までのサイトを抽出するか指定する
RANK = 20
pages_num = RANK + 1

AntisocialCount = 0

inputCellRow = 2
inputCellcolumn = 5

titleCounter = 0

writtenCellRow = 2
writtenCellcolumn = 6

#インプットデータ（Excel）を配置している絶対パスを設定する
wb = openpyxl.load_workbook( r'C:\Users' )

#シート取得
ws = wb[ "Sheet1" ]

#シートの行数を取得する。
maxRow = wb['Sheet1'].max_row

# 背景色を設定
fill = openpyxl.styles.PatternFill(patternType='solid',
                                   fgColor='ffd700', bgColor='ffd700')

for i in range( 2 , maxRow + 1 ):
    #Excelに入力された業者or顧客名を取得する
    
    if titleCounter == 0:
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn )
    
    #見出し語20個目なら行だけインクリメントしてインプット（代表者のインプット）
    if titleCounter == RANK:
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn + 1 )
    
    #見出し語40個目なら列をインクリメントして、インプット（次の会社名をインプット） 見出し語カウンターを0に戻す
    if titleCounter >=  RANK * 2:
        titleCounter = 0
        inputCellRow += 1
        writtenCellcolumn = 6
        cell = ws.cell( row = inputCellRow , column = inputCellcolumn )

    #検索ワードがなくなったら終了する
    if cell.value is None:
        print("検索ワードが無いので終了！")
        break
    
    #上で取得した業者or顧客名をstring型にキャストする。
    cellString = str( cell.value )

    #「㈱」と「㈲」を削除する（新聞記事は会社名を基本的に「㈱」と「㈲」を抜いて記載しており、新聞の記事を検索にヒットさせるため
    # 「医療法人」「株式会社」「(一社)」「合同会社」「土地家屋調査士法人」「(同)」「一般財団法人」「(有)」「合資会社」「土質試験協同組合」

    cellString_replaced1 = cellString.replace('㈱', '')
    cellString_replaced2 = cellString_replaced1.replace('㈲', '')
    cellString_replaced3 = cellString_replaced2.replace('医療法人', '')
    cellString_replaced4 = cellString_replaced3.replace('株式会社', '')
    cellString_replaced5 = cellString_replaced4.replace('(一社)', '')
    cellString_replaced6 = cellString_replaced5.replace('合同会社', '')
    cellString_replaced7 = cellString_replaced6.replace('土地家屋調査士法人', '')
    cellString_replaced8 = cellString_replaced7.replace('(同)', '')
    cellString_replaced9 = cellString_replaced8.replace('一般財団法人', '')
    cellString_replaced10 = cellString_replaced9.replace('(有)', '')
    cellString_replaced11 = cellString_replaced10.replace('合資会社', '')
    cellString_replaced12 = cellString_replaced11.replace('土質試験協同組合', '')

    # 「代表取締役」「院長」「総経理」「学長」「代表理事」「President」「最高経営責任者」「弁護士」「取締役社長」「Director」「所長」
    cellString_replaced13 = cellString_replaced12.replace('代表取締役', '')
    cellString_replaced14 = cellString_replaced13.replace('院長', '')
    cellString_replaced15 = cellString_replaced14.replace('総経理', '')
    cellString_replaced16 = cellString_replaced15.replace('学長', '')
    cellString_replaced17 = cellString_replaced16.replace('代表理事', '')
    cellString_replaced18 = cellString_replaced17.replace('President', '')
    cellString_replaced19 = cellString_replaced18.replace('最高経営責任者', '')
    cellString_replaced20 = cellString_replaced19.replace('弁護士', '')
    cellString_replaced21 = cellString_replaced20.replace('取締役社長', '')
    cellString_replaced22 = cellString_replaced21.replace('Director', '')
    cellString_replaced23 = cellString_replaced22.replace('所長', '')

    # 「代表社員」「工場長」「取締役」「支店長」「代表」「共同代表」「次長」「代表執行役」「代表者」「会長」
    cellString_replaced24 = cellString_replaced23.replace('代表社員', '')
    cellString_replaced25 = cellString_replaced24.replace('工場長', '')
    cellString_replaced26 = cellString_replaced25.replace('取締役', '')
    cellString_replaced27 = cellString_replaced26.replace('支店長', '')
    cellString_replaced28 = cellString_replaced27.replace('共同代表', '')
    cellString_replaced29 = cellString_replaced28.replace('次長', '')
    cellString_replaced30 = cellString_replaced29.replace('代表執行役', '')
    cellString_replaced31 = cellString_replaced30.replace('代表者', '')
    cellString_replaced32 = cellString_replaced31.replace('代表', '')
    cellString_replaced33 = cellString_replaced32.replace('会長', '')
    cellString_replaced34 = cellString_replaced33.replace('社長', '')
    cellString_replaced35 = cellString_replaced34.replace('CEO', '')
    cellString_replaced36 = cellString_replaced35.replace('/', '')
    cellString_replaced37 = cellString_replaced36.replace('兼', '')
    cellString_replaced38 = cellString_replaced37.replace('Partner', '')
    cellString_replaced39 = cellString_replaced38.replace('　', '')
    cellString_replaced40 = cellString_replaced39.replace(' ', '')
    
    # テスト用出力
    #ws.cell( row = i, column = 2  ).value = cellString_replaced40
    
    #業者or顧客名に反社ワードを連結させる。
    search_word = cellString_replaced40 + ANTISOCIAL＿FORCES 
    
    #ここを通る時、rは2に戻す。
    r = 5
    
    print( f'【検索ワード】{search_word}' )
    
    # Googleから検索結果ページを取得する
    url = f'https://www.google.co.jp/search?hl=ja&num={pages_num}&q={search_word}'
    request = requests.get( url )
    print(request)
    # Googleのページ解析を行う
    soup = BeautifulSoup( request.text, "html.parser" )
    search_site_list = soup.select( 'div.kCrYT > a' )

    # ページ解析と結果の出力
    for rank, site in zip(range( 1 , pages_num ) , search_site_list):
            #見出し語のカウント
            titleCounter += 1
            
            try:
                #site.select('h3.zBAuLc')[0]では、<h3 class="zBAuLc"/>サイトタイトル/</h3>部分を抽出しています。
                site_title = site.select( 'h3' )[0].text
            except IndexError:
                site_title = site.select( 'img' )[0]['alt']
                continue
            site_url = site['href'].replace( '/url?q=', '' )
            
            #余分な文字列（クエリパラメータ）を削除する。
            if '&' in site_url:
                print('URLから&を検出')
                txt = site_url
                anp = txt.find('&')
                site_url = txt[:anp] 
                print( 'site_url = ' + site_url )    

            if '%' in site_url:
                print('URLから%を検出')
                per = site_url.find('%')
                site_url = site_url[:per]     
                print( 'site_url = ' + site_url ) 
            
            if '?' in site_url:
                print('URLから?を検出')
                qes = site_url.find('?')
                site_url = site_url[:qes]     
                print( 'site_url = ' + site_url ) 

            print('最終的なURL = ' + site_url)
            
            # 結果を出力する
            print( str(rank) + "位: " + site_title )
            print( "!!!!!!!!!!!!!!!!!!!!!! titleCounter = " + str(titleCounter)  + " !!!!!!!!!!!!!!!!!!!!!! ")
            
            writtenCellcolumn += 1
            ws.cell( row = inputCellRow, column = writtenCellcolumn ).value = site_title
            ws.cell( row = inputCellRow, column = writtenCellcolumn ).hyperlink = site_url
    
            #ユーザエージェントを設定　「私は○○と言うブラウザです！」と自己紹介するためのコード。これが無いと止まる。
            try:
                headers_dic = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36"}
                res = requests.get( site_url, headers=headers_dic )
                
            #例外処理　プログラム内で対応可能なエラーをスルーする
            except  TooManyRedirects as e:
                print('catch TooManyRedirects:', e)
                r = r + 1
                if r == RANK + r:
                    r = 5
                continue
            except  ConnectionError as e:
                print('catch ConnectionError:', e)
                r = r + 1
                if r == RANK + r:
                    r = 5
                continue

            #取得した見出し語について、反社ワードを１つ１つ照合して、反社ワードが含まれる見出し語が書き込まれたセルをハイライトする。
            for x in range(0,127):
                if ANTISOCIAL_FORCES_LIST[x] in site_title: 
                    print('!!!!!!!!!!反社ワード走査!!!!!!!!!!!!' + ANTISOCIAL_FORCES_LIST[x])
                    ws.cell( row = inputCellRow, column = writtenCellcolumn ).fill = fill

            #インクリメントと調整
            r = r + 1
            if r == RANK + r:
                r = 5
            
            #ハッキングやサイバー攻撃（DOS攻撃）を疑われないためにあえて処理を遅く調節している。  
            #相手サーバに負担をかけないために、タイムスリープを設定（１秒以上取っておくと安全）            
            #一回ごとに保存した方がリスクが小さいかも知れない（動作が遅くなったりするかも？）
            wb.save( r'C:\Users\○○\Desktop\2024年度反社チェック\○○.xlsx' )
            SLEEP_TIME = random.randrange(2)

            time.sleep( SLEEP_TIME )


wb.save( r'C:\Users\○○\Desktop\2024年度反社チェック\○○.xlsx' )
# 時間計測終了
time_end = time.time()
# 経過時間（秒）
tim = time_end- time_sta

print( str( tim ) + "秒かかりました。")
wb.close()